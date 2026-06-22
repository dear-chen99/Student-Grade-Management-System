"""
Excel 处理模块 - Excel Handler Module.

提供 Excel 文件的导入导出功能，支持 .xlsx 和 .xls 格式。

本模块基于 ``openpyxl``（优先）和 ``xlrd``（降级）实现 Excel 的读写操作，
包含以下核心功能：

- ``create_template``: 生成包含表头和示例数据的导入模板。
- ``import_from_excel``: 从 Excel 批量导入学生基本信息与各科成绩，自动识别表头。
- ``export_to_excel``: 将现有成绩数据（含排名、总分、平均分、等级）导出为 Excel。
- ``get_default_filename``: 生成带日期的默认导出文件名。

依赖::

    建议安装 ``openpyxl`` 以获得完整功能::

        pip install openpyxl

Attributes:
    EX_OK: 布尔标志，指示 Excel 处理库是否可用。
"""

import datetime
import logging
import os
from typing import Any, Optional, Tuple

# ------------------------------------------------------------------
# 模块级日志配置
# ------------------------------------------------------------------
# 配置独立的日志记录器，便于在导入/导出过程中追踪操作细节和异常。
logger = logging.getLogger("ExcelHandler")
logger.setLevel(logging.INFO)
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(
        logging.Formatter(
            "[%(asctime)s] %(levelname)s - %(name)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    logger.addHandler(_handler)


# ------------------------------------------------------------------
# 自定义异常类
# ------------------------------------------------------------------


class ExcelHandlerError(Exception):
    """Excel 处理模块基础异常类.

    所有本模块自定义异常的基类，便于调用方统一捕获。

    Attributes:
        message: 错误描述信息。
    """

    def __init__(self, message: str = "Excel 处理错误") -> None:
        """初始化异常实例.

        Args:
            message: 错误描述信息。
        """
        super().__init__(message)
        self.message = message


class ExcelLibraryNotFoundError(ExcelHandlerError):
    """Excel 库未安装错误.

    当系统既未安装 ``openpyxl`` 也未安装 ``xlrd`` 时抛出，
    提示用户安装必要的第三方库。
    """

    def __init__(self) -> None:
        """初始化 Excel 库未安装错误."""
        super().__init__("请执行：pip install openpyxl 安装 Excel 处理库")


class ExcelReadError(ExcelHandlerError):
    """Excel 文件读取错误.

    在打开或读取 Excel 文件过程中发生 IO 错误、格式错误等时抛出。

    Attributes:
        filepath: 发生错误的文件路径。
        message: 包含路径和原因的错误描述。
    """

    def __init__(self, filepath: str, reason: str = "") -> None:
        """初始化 Excel 读取错误，记录文件路径和原因.

        Args:
            filepath: 目标文件路径。
            reason: 失败原因补充说明。
        """
        msg = f"无法读取 Excel 文件: {filepath}"
        if reason:
            msg += f" ({reason})"
        super().__init__(msg)


class ExcelWriteError(ExcelHandlerError):
    """Excel 文件写入错误.

    在保存 Excel 文件过程中发生权限不足、路径不存在等错误时抛出。

    Attributes:
        filepath: 发生错误的文件路径。
        message: 包含路径和原因的错误描述。
    """

    def __init__(self, filepath: str, reason: str = "") -> None:
        """初始化 Excel 写入错误，记录文件路径和原因.

        Args:
            filepath: 目标文件路径。
            reason: 失败原因补充说明。
        """
        msg = f"无法写入 Excel 文件: {filepath}"
        if reason:
            msg += f" ({reason})"
        super().__init__(msg)


class ExcelFormatError(ExcelHandlerError):
    """Excel 文件格式错误.

    当文件扩展名不符合预期或内容不是有效的 Excel 格式时抛出。

    Attributes:
        filepath: 发生错误的文件路径。
        message: 包含路径和原因的错误描述。
    """

    def __init__(self, filepath: str, reason: str = "") -> None:
        """初始化 Excel 格式错误，记录文件路径和原因.

        Args:
            filepath: 目标文件路径。
            reason: 失败原因补充说明。
        """
        msg = f"Excel 文件格式不正确: {filepath}"
        if reason:
            msg += f" ({reason})"
        super().__init__(msg)


class ExcelEmptyError(ExcelHandlerError):
    """Excel 文件内容为空错误.

    当读取到的 Excel 文件只有表头或完全空白时抛出。

    Attributes:
        message: 错误描述信息。
    """

    def __init__(self, filepath: str = "") -> None:
        """初始化 Excel 内容为空错误，记录文件路径.

        Args:
            filepath: 目标文件路径，可为空。
        """
        base = f"文件内容为空: {filepath}" if filepath else "文件内容为空"
        super().__init__(base)


# ------------------------------------------------------------------
# Excel 库可用性检查（分离读写依赖，防崩溃）
# ------------------------------------------------------------------
OPENPYXL_OK = False
XLDR_OK = False
try:
    import openpyxl  # noqa: F401

    OPENPYXL_OK = True
    logger.debug("openpyxl 导入成功")
except ImportError:
    pass

try:
    import xlrd  # noqa: F401

    XLDR_OK = True
    logger.debug("xlrd 导入成功")
except ImportError:
    pass

# 只要有一个能读，就算可用（用于 import_from_excel）
EX_OK = OPENPYXL_OK or XLDR_OK

if not EX_OK:
    logger.warning("未检测到 Excel 处理库，请执行: pip install openpyxl")


def is_excel_available() -> bool:
    """检查 Excel 处理库是否可用.

    Returns:
        True 如果 openpyxl 或 xlrd 至少有一个可用。
    """
    return EX_OK


def _validate_filepath(filepath: str) -> str:
    """校验文件路径的合法性.

    对路径进行非空检查并去除首尾空白字符。

    Args:
        filepath: 文件路径字符串。

    Returns:
        规范化后的文件路径（去除首尾空白）。

    Raises:
        ExcelHandlerError: 路径为空或不是有效字符串时抛出。
    """
    if not filepath or not isinstance(filepath, str):
        raise ExcelHandlerError("文件路径不能为空")
    filepath = filepath.strip()
    if not filepath:
        raise ExcelHandlerError("文件路径不能为空")
    return filepath


# ------------------------------------------------------------------
# 公共方法
# ------------------------------------------------------------------


def create_template(filepath: str, subjects: list[str]) -> bool:
    """创建 Excel 成绩导入模板.

    生成包含表头（学号、姓名、班级 + 科目）和一行示例数据的 .xlsx 文件，
    方便用户按规范录入成绩后批量导入系统。

    Args:
        filepath: 模板保存路径（建议以 .xlsx 结尾）。
        subjects: 科目名称列表，将依次作为表头列。

    Returns:
        True 表示创建成功，False 表示失败（如库未安装、路径无效等）。
    """
    if not OPENPYXL_OK:
        logger.error("创建模板失败: openpyxl 未安装")
        return False

    try:
        filepath = _validate_filepath(filepath)
    except ExcelHandlerError as e:
        logger.error("创建模板失败: %s", e)
        return False

    if not subjects:
        logger.warning("创建模板时科目列表为空")

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            logger.error("创建模板失败: 无法获取活动工作表")
            return False
        ws.title = "模板"
        # 写入表头与示例数据行
        ws.append(["学号", "姓名", "班级"] + list(subjects))
        ws.append(["2024001", "张三", "一班"] + [""] * len(subjects))
        wb.save(filepath)
        logger.info("Excel 模板已创建: %s", filepath)
        return True

    except OSError as e:
        logger.error("创建模板写入失败: %s - %s", filepath, e)
        return False
    except Exception as e:
        logger.error("创建模板未知错误: %s", e)
        return False


def import_from_excel(filepath: str, data_manager: Any) -> Tuple[int, Optional[str]]:
    """从 Excel 文件导入学生成绩数据.

    自动识别 .xlsx 和 .xls 格式，解析表头中的学号、姓名、班级及各科目列，
    将有效成绩逐行导入到 ``DataManager`` 中。对于不存在的科目会自动创建，
    已存在的学生会更新基本信息并追加成绩。

    解析规则::

        1. 表头必须包含 "学号" 和 "姓名" 列。
        2. 若存在 "班级" 列则按单元格值导入；否则所有学生归入默认班级 "24计软技师"。
        3. 除学号、姓名、班级、总分、总学分外的列均视为科目。
        4. 成绩为空、"None" 或超出 0-100 范围的值将被跳过并记录警告日志。
        5. 旧格式 1-3 位纯数字学号会自动转换为 2024xxxx 格式。

    Args:
        filepath: Excel 文件路径。
        data_manager: DataManager 实例，用于写入学生与成绩数据。

    Returns:
        二元组 ``(imported_count, error_message)``。
        ``imported_count`` 为成功导入的学生人数；
        ``error_message`` 为错误描述字符串，成功时为 ``None``。
    """
    if not EX_OK:
        msg = "请执行：pip install openpyxl"
        logger.error(msg)
        return (0, msg)

    try:
        filepath = _validate_filepath(filepath)
    except ExcelHandlerError as e:
        return (0, str(e))

    if not os.path.exists(filepath):
        msg = f"文件不存在: {filepath}"
        logger.error(msg)
        return (0, msg)

    try:
        # 根据扩展名选择对应的读取库
        if filepath.lower().endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                ws = wb.active
                if ws is None:
                    return (0, "无法读取工作表")
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                logger.info("读取 .xlsx 文件: %s (%d 行)", filepath, len(rows))
            except OSError as e:
                logger.error("读取 .xlsx 文件失败: %s - %s", filepath, e)
                return (0, f"无法读取文件: {e}")
            except Exception as e:
                logger.error("解析 .xlsx 文件失败: %s - %s", filepath, e)
                return (0, f"文件格式不正确: {e}")
        else:
            try:
                rb = xlrd.open_workbook(filepath)
                sh = rb.sheet_by_index(0)
                # 将每一行转换为字符串列表，空单元格统一为空字符串
                rows = [
                    [str(c).strip() if c else "" for c in sh.row_values(i)]
                    for i in range(sh.nrows)
                ]
                logger.info("读取 .xls 文件: %s (%d 行)", filepath, len(rows))
            except OSError as e:
                logger.error("读取 .xls 文件失败: %s - %s", filepath, e)
                return (0, f"无法读取文件: {e}")
            except Exception as e:
                logger.error("解析 .xls 文件失败: %s - %s", filepath, e)
                return (0, f"文件格式不正确: {e}")

        # 至少要有表头 + 一行数据
        if len(rows) < 2:
            logger.warning("Excel 文件内容为空: %s", filepath)
            return (0, "文件内容为空，至少需要表头和数据行")

        # 解析表头：统一转为字符串并去除空白
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        if not any(header):
            return (0, "表头为空，无法识别列")

        # 定位学号列和姓名列（默认第0列和第1列，若表头含关键字则按关键字匹配）
        id_idx = next((i for i, h in enumerate(header) if "学号" in str(h)), 0)
        name_idx = next((i for i, h in enumerate(header) if "姓名" in str(h)), 1)

        # 明确跳过列：只跳过学号和姓名
        skip_idxs = {id_idx, name_idx}

        # 处理班级列
        if "班级" in header:
            class_idx = next((i for i, h in enumerate(header) if "班级" in str(h)), 2)
            skip_idxs.add(class_idx)
            default_class_name = None
        else:
            # 没有班级列，使用默认班级
            default_class_name = "24计软技师"
            logger.info("未找到班级列，所有学生将自动归入: %s", default_class_name)

        # 解析科目列：除跳过列外，且列名非总分/总学分的都视为科目
        subject_idx: dict[str, int] = {}
        for i, h in enumerate(header):
            if i not in skip_idxs and h and h not in ["总分", "总学分"]:
                if h not in data_manager.subjects:
                    try:
                        data_manager.add_subject(h)
                        logger.info("自动添加科目: %s", h)
                    except Exception as e:
                        logger.warning("自动添加科目失败: %s - %s", h, e)
                        continue
                subject_idx[h] = i

        if not subject_idx:
            logger.warning("未找到科目列: %s", filepath)
            return (0, "未找到科目列，请检查表头是否包含科目名称")

        # 逐行导入数据
        imported = 0
        import_errors: list[str] = []
        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                values = [c for c in row]
                # 跳过全空行
                if all(not str(c).strip() if c is not None else True for c in values):
                    continue

                student_id = (
                    str(values[id_idx]).strip()
                    if id_idx < len(values) and values[id_idx] is not None
                    else ""
                )
                if not student_id:
                    continue

                # 旧格式 1-3 位纯数字 → 2024xxxx
                if student_id.isdigit() and 1 <= len(student_id) <= 3:
                    student_id = f"2024{student_id.zfill(3)}"

                name = (
                    str(values[name_idx]).strip()
                    if name_idx < len(values) and values[name_idx] is not None
                    else ""
                )

                # 获取班级名称
                if default_class_name is not None:
                    class_name = default_class_name
                else:
                    class_name = (
                        str(values[class_idx]).strip()
                        if class_idx < len(values) and values[class_idx] is not None
                        else ""
                    )

                # 收集该行的各科成绩
                scores: dict[str, float] = {}
                for subject, idx in subject_idx.items():
                    value = (
                        str(values[idx]).strip()
                        if idx < len(values) and values[idx] is not None
                        else ""
                    )
                    if value and value != "None":
                        try:
                            parsed = float(value)
                            if 0 <= parsed <= 100:
                                scores[subject] = parsed
                            else:
                                logger.warning(
                                    "第 %d 行 %s 成绩超出范围: %s",
                                    row_idx,
                                    subject,
                                    parsed,
                                )
                        except ValueError:
                            logger.warning(
                                "第 %d 行 %s 成绩无效: %s", row_idx, subject, value
                            )

                # 写入 DataManager：已存在则更新，不存在则新增
                try:
                    if data_manager.exists(student_id):
                        data_manager.update_student(student_id, name, class_name)
                    else:
                        data_manager.add_student(student_id, name, class_name)
                    data_manager.batch_set(student_id, scores)
                    imported += 1
                except Exception as e:
                    msg = f"第 {row_idx} 行导入失败: {e}"
                    import_errors.append(msg)
                    logger.warning(msg)

            except Exception as e:
                msg = f"第 {row_idx} 行解析失败: {e}"
                import_errors.append(msg)
                logger.warning(msg)

        if import_errors:
            logger.warning(
                "导入完成，共 %d 条成功，%d 条失败", imported, len(import_errors)
            )

        return (imported, None)

    except Exception as e:
        logger.error("导入 Excel 文件异常: %s - %s", filepath, e)
        return (0, f"导入失败: {e}")


def export_to_excel(filepath: str, data_manager: Any) -> bool:
    """导出成绩数据到 Excel 文件.

    包含排名、学号、姓名、班级、各科成绩、总分、平均分和等级（优秀/良好/及格/不及格）。
    等级根据平均分划分::

        >= 90 -> 优秀
        >= 75 -> 良好
        >= 60 -> 及格
        <  60 -> 不及格

    Args:
        filepath: 保存路径（建议以 .xlsx 结尾）。
        data_manager: DataManager 实例，需提供 ``subjects`` 属性和 ``ranking()`` 方法。

    Returns:
        True 表示导出成功，False 表示失败。
    """
    if not OPENPYXL_OK:
        logger.error("导出 Excel 失败: openpyxl 未安装")
        return False

    try:
        filepath = _validate_filepath(filepath)
    except ExcelHandlerError as e:
        logger.error("导出 Excel 失败: %s", e)
        return False

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            logger.error("导出 Excel 失败: 无法获取活动工作表")
            return False
        ws.title = "成绩表"

        subjects = (
            list(data_manager.subjects) if hasattr(data_manager, "subjects") else []
        )
        ws.append(
            ["排名", "学号", "姓名", "班级"] + subjects + ["总分", "平均分", "等级"]
        )

        # 逐行写入每个学生的成绩与排名
        ranking = data_manager.ranking() if hasattr(data_manager, "ranking") else []
        for rank_data in ranking:
            avg = rank_data.get("avg", 0)
            # 根据平均分判定等级
            if avg >= 90:
                level = "优秀"
            elif avg >= 75:
                level = "良好"
            elif avg >= 60:
                level = "及格"
            else:
                level = "不及格"

            row_data = (
                [
                    rank_data.get("rank", ""),
                    rank_data.get("id", ""),
                    rank_data.get("name", ""),
                    rank_data.get("class", ""),
                ]
                + [rank_data.get("scores", {}).get(s, "-") for s in subjects]
                + [rank_data.get("total", 0), rank_data.get("avg", 0), level]
            )
            ws.append(row_data)

        wb.save(filepath)
        logger.info("Excel 导出成功: %s", filepath)
        return True

    except OSError as e:
        logger.error("导出 Excel 写入失败: %s - %s", filepath, e)
        return False
    except Exception as e:
        logger.error("导出 Excel 未知错误: %s", e)
        return False


def get_default_filename(extension: str = "xlsx") -> str:
    """获取默认导出文件名（含日期）.

    Args:
        extension: 文件扩展名，默认 "xlsx"（可传入 "xls" 等）。

    Returns:
        格式为 "成绩_YYYY-MM-DD.{extension}" 的文件名。
    """
    ext = extension.lstrip(".")
    return f"成绩_{datetime.date.today()}.{ext}"
